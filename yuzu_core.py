# -*- coding: utf-8 -*-
"""
在庫融通 マッチング共通エンジン（yuzu_core.py）

【何をするスクリプトか】
店舗間 在庫融通ツールの「頭脳」だけを1ファイルにまとめた共通エンジンです。
  ・薬VANの在庫マスタ（.csv / .xls / .xlsx）を「バイト＋ファイル名」で読み込む道具
  ・過剰／不足の判定、引取候補店（①不足中→②使用中→③参考:過剰）の組み立て
  ・全店ぶんを一括計算する compute_matching(stores)（4シート相当データ＋自己検算を返す）
  ・Excel出力 write_excel()
を提供します。

  【2026-07-27 追加】④受け手ビュー用に、引取候補店を「1件1行」で持つ丸め無しの構造化データを追加しました。
    build_candidate_entries()（丸めを通さず候補を全件返す）＋ compute_matching の戻り値 'candidate_rows'。
    既存キー・Excel4シート・Gシート出力は一切変えていません（増やすだけ）。

  【2026-07-28 追加】CSVの文字コードを自動判定するようにしました（UTF-8 →Shift-JIS の順）。
    薬VANのCSVは店によって UTF-8(BOM付き) と Shift-JIS(cp932) の2種類が出力されるためです。
  【2026-07-28 追加】②③の画面用に、引取候補店を「店名だけ」つないだ build_candidate_names() と、
    compute_matching の提案行キー '引取候補店（店名のみ）' を追加しました。
    既存の『引取候補店』（tier・消化目安つき＝Excel／Gシート用）は一切変えていません（増やすだけ）。

このファイルは「単一の真実」です。次の2つが必ず同じ計算をするよう、両者ともここを呼びます。
  ・コマンド版：create_yuzu_list.py（input フォルダのファイルをパスから読む）
  ・アプリ版　：streamlit_app.py（店がアップロードしたファイルのバイトをそのまま読む）
ファイル読み込みを「バイト＋ファイル名」に一般化してあるので、
コマンド版はパスからバイトを読んで、アプリ版はアップロード物のバイトを、同じ関数に渡せます。

【必要なライブラリのインストール（コマンドプロンプトで実行）】
    pip install openpyxl xlrd==2.0.1
    ※ xlrd は薬VANが .xls（古いExcel＝BIFF/OLE2形式）で出力したファイルを読むために使います。
      xlrd 2系は .xls 専用（.xlsx は読みません）。.xlsx は openpyxl で読みます。
    ※ .csv だけしか使わないなら xlrd は無くても動きます。

※ venv不要。Windows専用パス。コメント・ログ・エラーメッセージはすべて日本語です。
"""

import os
import csv
import io
import datetime
from collections import defaultdict, Counter

# openpyxl（Excel出力・.xlsx読み込み）
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# 設定（CONFIG）… 抽出条件・列名はここに集約。ロジック本体に値を直書きしない。
#   ※ 入力／出力フォルダやGシート設定ファイルのパスは、呼び出し側（コマンド版）が持ちます。
#     ここには「計算のルール」だけを置きます。
# ============================================================================
CONFIG = {
    # --- 前処理の除外（True で除外を有効化）---
    'exclude_deleted':   True,   # 削除フラグ = TRUE の行を除外
    'exclude_untreated': True,   # 取扱フラグ = FALSE の行を除外
    'exclude_otc':       True,   # OTCフラグ = TRUE の行を除外

    # --- キー列（店舗間の突合キー。位置ではなく列名で特定する）---
    'col_key':          '個別医薬品CD',   # 第一キー（YJコード）
    'col_key_fallback': 'レセプト電算CD', # キーが空のときの代用

    # --- 直近6ヶ月の出庫（当月は未確定なので含めず、前月〜六ヶ月前の6つを使う）---
    # ★重要★ 薬VANの「出庫数」は、実際に調剤して払い出すと【マイナス（負値）】で記録されます。
    #   （例：アリピプラゾール錠６ｍｇ「ＹＤ」は6ヶ月の出庫数合計 -2778 ＝バリバリ使用中）
    #   一方「出庫回数」は払い出した回数で【プラス（正値）】です。
    #   → 「使っているか？」の判定は符号に左右されない“出庫回数合計>0”で行い、
    #      月あたりの消化量は“出庫数合計の絶対値÷6”で求めます（下の usage_* 関数・pace 参照）。
    'outflow_qty_cols_6m': ['前月出庫数', '前々月出庫数', '前々々月出庫数',
                            '四ヶ月前出庫数', '五ヶ月前出庫数', '六ヶ月前出庫数'],
    'outflow_cnt_cols_6m': ['前月出庫回数', '前々月出庫回数', '前々々月出庫回数',
                            '四ヶ月前出庫回数', '五ヶ月前出庫回数', '六ヶ月前出庫回数'],

    # --- 引取候補の3段階（tier）判定 ---
    #   ① 不足中          ：受け取り側が 安全在庫数>0 かつ 在庫数<安全在庫数（is_shortage）
    #   ② 使用中(適正)    ：不足中でなく、6ヶ月出庫回数合計>0 で、かつ受け取り側がその品目を
    #                        「過剰保有」していない（＝ここに送ると有効活用される本命候補）
    #   ③ 参考:使用中(過剰)：6ヶ月出庫回数合計>0 だが、受け取り側も「過剰保有」している
    #                        （＝送っても余りを増やすだけなので“参考”として下位・別枠に表示）
    #
    # 「過剰保有（holds_excess）」の定義 ＝ 受け取り側がすでにその品目を余らせているか：
    #   過剰在庫区分が非空 または 不動区分が非空 または 過剰数>0
    #   ※ 期限切迫は「使えば消える」ので過剰保有には含めない
    #     （出し手判定 is_supplier＝デッド／期限切迫 とは別定義であることに注意。
    #       holds_excess は“受け取り側が余らせているか”の判定なので、A案でも変更しない）
    'holds_excess_flag_cols': ['過剰在庫区分', '不動区分'],  # このどれかが非空なら過剰保有
    'holds_excess_qty_col':   '過剰数',                     # この数量が >0 なら過剰保有
    # tierごとの表示ラベル
    'tier_labels': {'short': '不足中', 'use_ok': '使用中', 'use_ref': '参考:過剰だが使用中'},

    # --- デッド（不動）の判定：不動区分のどの値をデッドに含めるか ---
    # ★薬VANの「不動区分」は、最終出庫日からどれだけ経ったかで3段階に分かれます
    #   （2026-07-29に実データ2店＝さと和光・東立石で解読。それまで意味は未確定でした）：
    #     B ＝ 最終出庫から 3ヶ月以上 6ヶ月未満
    #     O ＝ 最終出庫から 6ヶ月以上 12ヶ月未満
    #     R ＝ 最終出庫から 12ヶ月以上（薬VANが持つデータは最長24ヶ月）
    # ★本間部長判断（2026-07-29・A案）＝デッドは「6ヶ月以上動いていない品」に絞る。
    #   → O と R だけをデッドとし、B（3〜6ヶ月）はデッドに含めない。
    # ※ 旧仕様（3ヶ月以上＝不動区分が非空なら全部デッド）に戻すときは、ここを [] （空リスト）に。
    # ※「Oと Rを載せる」ではなく「Bを外す」書き方にしているのは、薬VANが将来 B/O/R 以外の値を
    #   返してもデッドとして残り、黙って取りこぼさないようにするためです。
    'dead_exclude_kubun_values': ['B'],

    # --- 区分の扱い ---
    # 過剰/不動/期限切迫/発注候補：空欄＝非該当、それ以外（R/B/O 等）＝該当
    # 法規制（麻薬/覚醒剤/向精神薬/毒薬/劇薬）：'0' か空欄＝非該当、それ以外＝該当
    #   ※ 実データでは法規制列は「該当しない」を '0' で表す（空欄ではない）ため区別が必要
    'legal_not_flagged_values': ['', '0'],

    # 融通提案から完全に除外する法規制列（麻薬・覚醒剤）
    'legal_exclude_cols': ['麻薬区分', '覚醒剤区分'],
    # 除外はしないが「要記録」警告を出す法規制列（列名 → 表示ラベル）
    'legal_warn_cols': {'向精神薬区分': '向精神薬', '毒薬区分': '毒薬', '劇薬区分': '劇薬'},

    # 融通提案に載せる最低の在庫金額（円）。これ未満の少額品は載せない。
    #   （本間部長判断 2026-07-27：少額品まで並ぶと本当に動かすべき品が埋もれるため）
    'min_supply_amount': 1500,

    # 引取候補店の最大表示数（あふれた分は「他N店」と表示）
    'max_candidates': 5,
    # 有効期限「1年以内」を黄色でハイライトする閾値（月）
    'expiry_yellow_months': 12,
    # 「期限切迫」として融通提案に載せる（＝赤で目立たせる）閾値（月）。
    #   有効期限が基準日から この月数以内 の品だけを期限切迫として扱う（既定＝5ヶ月以内）。
    #   ★2026-08-04（本間部長・A案）：従来は「期限切迫区分が非空なら全部」で、期限が
    #     半年〜1年先の“まだ動いている品”まで載っていた。実データで薬VANの期限切迫区分は
    #     B＝期限まで6〜11ヶ月先／O＝3〜5ヶ月／R＝3ヶ月未満（期限切れ含む）と判明したため、
    #     区分に頼らず「有効期限までの実日数」で判定する方式に変えた（店ごとの基準日を使う）。
    #   ※もっと絞る／広げるときは、この数字を 3 や 4 に変えるだけでよい。
    'expiry_within_months': 5,

    # 失効予約（当月＞受取予定月＝もう受け取れない予約）を _予約 タブから自動で消すか。
    #   既定＝False（消さない）。オフの間は「画面内の有効判定だけで無害化」する
    #   ＝失効予約は表示・マッチングに一切効かないが、_予約 タブには残しておく。
    #   ★True にした場合でも、掃除専用の書き込みは作らない＝予約／取消の“保存に相乗り”して
    #     落とす（Google APIの呼び出し回数を増やさないため。plan_reservations / cancel_reservations 参照）。
    'reservation_autoclean': False,
}

# 必須列（このどれかが欠けている店はスキップ）
REQUIRED_COLS = ['個別医薬品CD', 'レセプト電算CD', '薬品名', '単位', 'メーカ名',
                 '在庫数', '全在庫数', '安全在庫数', '過剰数', '過剰数金額', '薬価金額',
                 '過剰在庫区分', '不動区分', '期限切迫区分', '有効期限', 'ロットNO',
                 '最終出庫日', '削除フラグ', '取扱フラグ', 'OTCフラグ']

# 区分値の分布ログに出す列
DIST_COLS = ['過剰在庫区分', '不動区分', '期限切迫区分', '発注候補区分']

# アプリ版でGシート保管庫へ入れる前に残す「マッチングに要る約35列」。
#   薬VANは243〜300列と巨大で、Gシート（1ブック1000万セル上限）を圧迫するため、
#   エンジンが実際に使う列だけへ圧縮してから保管する（下の slim_row / slim_rows 参照）。
KEEP_COLS = [
    # 突合キー・基本情報
    '個別医薬品CD', 'レセプト電算CD', '薬品名', '単位', 'メーカ名',
    # 在庫・過剰・安全在庫（薬価/薬価金額＝在庫金額を出すのに使う）
    '在庫数', '全在庫数', '安全在庫数', '過剰数', '過剰数金額', '薬価', '薬価金額',
    # 区分（過剰・不動・期限切迫・発注候補）
    '過剰在庫区分', '不動区分', '期限切迫区分', '発注候補区分',
    # 期限・ロット・最終出庫
    '有効期限', 'ロットNO', '最終出庫日',
    # 前処理フラグ（削除・取扱・OTC）
    '削除フラグ', '取扱フラグ', 'OTCフラグ',
    # 法規制区分（除外・警告）
    '麻薬区分', '覚醒剤区分', '向精神薬区分', '毒薬区分', '劇薬区分',
    # 直近6ヶ月の出庫数（負値＝払い出し）
    '前月出庫数', '前々月出庫数', '前々々月出庫数',
    '四ヶ月前出庫数', '五ヶ月前出庫数', '六ヶ月前出庫数',
    # 直近6ヶ月の出庫回数（正値＝払い出し回数）
    '前月出庫回数', '前々月出庫回数', '前々々月出庫回数',
    '四ヶ月前出庫回数', '五ヶ月前出庫回数', '六ヶ月前出庫回数',
]


# ============================================================================
# 小さな道具（数値・日付・区分の判定）
# ============================================================================
def parse_num(s):
    """ "1,428.00" のような桁区切りカンマ＋文字列を float にする。空欄・変換不可は 0.0 """
    if s is None:
        return 0.0
    s = str(s).strip().replace(',', '')
    if s == '':
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(s):
    """ 'YYYY/M/D' 等の日付文字列を date にする。読めなければ None """
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y/%m', '%Y%m%d'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def fmt_date(d):
    return d.strftime('%Y/%m/%d') if d else ''


def fmt_ym(d):
    return d.strftime('%Y/%m') if d else '期限不明'


def month_diff(a, b):
    """ a から b までのおおよその月数（bが未来なら正） """
    return (b.year - a.year) * 12 + (b.month - a.month) + (b.day - a.day) / 30.0


def fmt_months(m):
    """ 消化目安の月数を読みやすく（小数1桁、整数なら整数で） """
    m = round(m, 1)
    return str(int(m)) if m == int(m) else str(m)


# --- 年月（YYYYMM文字列）の計算。予約の「受取予定月」で使う（2026-08-01 追加）---
#   ★日付(date)ではなく年月(YYYYMM6桁の文字列)で扱うのは、予約の受取時期が「◯ヶ月後」という
#     月単位の約束であり、対象年月・受取予定月ともに6桁文字列で保管しているため。
def ym_offset(base_ym, target_ym):
    """ base_ym から target_ym までの月数（target が先なら正）。どちらも 'YYYYMM' の6桁文字列。
        読めない値が来たら 0 を返す（画面が落ちないように）。 """
    try:
        by, bm = int(str(base_ym)[:4]), int(str(base_ym)[4:6])
        ty, tm = int(str(target_ym)[:4]), int(str(target_ym)[4:6])
    except (ValueError, TypeError):
        return 0
    return (ty - by) * 12 + (tm - bm)


def ym_add(base_ym, months):
    """ 'YYYYMM' に months ヶ月を足した 'YYYYMM' を返す。読めない値ならそのまま返す。 """
    try:
        y, m = int(str(base_ym)[:4]), int(str(base_ym)[4:6])
    except (ValueError, TypeError):
        return str(base_ym or '')
    total = y * 12 + (m - 1) + int(months)
    return '%04d%02d' % (total // 12, total % 12 + 1)


def pickup_label(offset):
    """ 受取までの月数（0=今すぐ）を画面に出す短い言葉にする。0以下は『今すぐ』、それ以外は『Nヶ月後』。 """
    try:
        offset = int(offset)
    except (ValueError, TypeError):
        offset = 0
    return '今すぐ' if offset <= 0 else '%dヶ月後' % offset


def fmt_qty(x):
    """ 数量を読みやすく（整数なら整数、端数があれば小数2桁） """
    if x == int(x):
        return str(int(x))
    return str(round(x, 2))


def is_text_flag(v):
    """ 過剰/不動/期限切迫など：空欄でなければ「該当」 """
    return bool(str(v).strip())


def is_legal_flag(v):
    """ 法規制区分：'0' か空欄なら非該当、それ以外なら「該当」 """
    return str(v).strip() not in CONFIG['legal_not_flagged_values']


def g(row, col):
    """ 行から列を安全に取り出す（無ければ空文字） """
    return (row.get(col) or '').strip() if isinstance(row.get(col), str) else (row.get(col) or '')


# ============================================================================
# 行レベルの判定
# ============================================================================
# ★出し手（融通候補）の定義（A案・本間部長承認 2026-07-27）★
#   従来は「過剰在庫区分/不動区分/期限切迫区分のいずれか非空、または過剰数>0」（＝is_overstock）
#   でしたが、純粋な過剰（デッドでも期限切迫でもない、ただ在庫が多いだけの品）を提案から外し、
#   本当に動かすべき「デッド（不動）」と「期限切迫」だけを出し手とすることにしました。
#     ・is_dead(row)   … 不動区分がデッド対象の値（＝デッド在庫）
#                        ★2026-07-29（本間部長・A案）に「不動区分が非空なら全部」から
#                          「6ヶ月以上動いていない品（O/R）だけ」へ変更。
#                          B（3〜6ヶ月）はまだ動く見込みがあるためデッドに含めない。
#                          → CONFIG['dead_exclude_kubun_values'] を参照。
#     ・is_expiry(row, base_date) … 有効期限が基準日から CONFIG['expiry_within_months']（既定5）
#                        ヶ月以内なら期限切迫。有効期限が空欄・読めない品は期限切迫にしない（落とす）。
#                        ★2026-08-04（本間部長・A案）：従来は「期限切迫区分が非空なら全部」だったが、
#                          期限が半年〜1年先の“まだ動いている品”まで載っていたため、区分ではなく
#                          「有効期限までの実日数」で判定する方式に変更。基準日は店ごとに違い得るので
#                          モジュール共通変数にはせず、引数 base_date で受け取る（誤判定を防ぐため）。
#     ・is_supplier(row, base_date)＝ is_dead または is_expiry （＝新しい出し手）
#   カテゴリは排他・デッド優先：デッドなら「デッド」、デッドでなく期限切迫なら「期限切迫」。
#   （デッドかつ期限切迫の品は「デッド」に入れる。ただし期限切迫区分の値は各行に残し、
#     Excelでは赤ハイライト（＝有効期限まで5ヶ月以内）を維持する。）
def is_dead(row):
    """ デッド（不動）在庫か。
        不動区分が非空で、かつ CONFIG['dead_exclude_kubun_values'] に載っていない値なら True。
        既定＝O（6〜12ヶ月）と R（12ヶ月以上）だけがデッド。B（3〜6ヶ月）は含めない。
        ※ B/O/R 以外の見慣れない値が来たときはデッド側に残す（取りこぼし防止）。 """
    v = str(g(row, '不動区分')).strip()
    if not v:
        return False
    ng = [str(x).strip().upper() for x in CONFIG['dead_exclude_kubun_values']]
    return v.upper() not in ng


def is_expiry(row, base_date):
    """ 期限切迫か：有効期限が基準日から CONFIG['expiry_within_months']（既定5）ヶ月以内なら True。
        有効期限が空欄・読めない品は False（＝期限切迫にしない＝提案に載せない）。
        ★基準日（base_date）は店ごとに違い得るので引数で受け取る（A案・2026-08-04）。
        ※薬VANの期限切迫区分（B/O/R）には頼らない：区分ではなく「有効期限までの実日数」で判定する。
          （実データでは B＝6〜11ヶ月先／O＝3〜5ヶ月／R＝3ヶ月未満だが、将来値が変わっても
            日付判定なら影響を受けないため。） """
    exp = parse_date(g(row, '有効期限'))
    if exp is None:
        return False
    return month_diff(base_date, exp) <= CONFIG['expiry_within_months']


def is_supplier(row, base_date):
    """ 出し手（融通候補）か：デッド または 期限切迫 なら True。
        ※ 純粋な過剰（過剰在庫区分だけ／過剰数>0だけで、不動でも期限切迫でもない品）は
          出し手に含めない（A案）。 """
    return is_dead(row) or is_expiry(row, base_date)


def supplier_category(row, base_date):
    """ 出し手のカテゴリを返す（排他・デッド優先）：
        デッドなら 'デッド'、デッドでなく期限切迫なら '期限切迫'、どちらでもなければ ''（＝出し手でない）。 """
    if is_dead(row):
        return 'デッド'
    if is_expiry(row, base_date):
        return '期限切迫'
    return ''


def is_below_min_amount(row):
    """ 在庫金額が少額（既定 1,500円未満）で、融通提案に載せない品か。
        少額品まで並ぶと、本当に動かすべき品が埋もれてしまうため。
        しきい値は CONFIG['min_supply_amount'] で変えられる。 """
    return stock_amount(row) < CONFIG['min_supply_amount']


def exclusion_key(row):
    """ 除外リスト（店が「この品は出さない」と外した品目）で1品を特定するキー。
        個別医薬品CDが取れればそれを使い、取れない品だけ『名:薬品名』で代用する。
        ※ 保存側（画面）と判定側（ここ）で必ず同じ関数を使うこと。 """
    key, _ = row_key(row)
    return key or ('名:' + g(row, '薬品名'))


def stock_qty(row):
    """ 表示・提案に使う数量＝『在庫数』（その店がいま持っている全量）。
        ※ 旧仕様は『過剰数』（安全在庫を超えた分）でしたが、デッド（不動）品は
          在庫まるごとが動かす対象なので、在庫数に統一しました（本間部長判断 2026-07-27）。
          実データでは304件中141件で 過剰数≠在庫数、過剰数0（＝金額0で最下位に沈む）の
          デッド品も37件あり、在庫数のほうが実態に合います。 """
    return parse_num(g(row, '在庫数'))


def stock_amount(row):
    """ 表示・提案に使う金額＝『在庫金額』＝ 在庫数 × 薬価。
        薬VANの『薬価金額』列がまさにこの値（実データ1,431行中1,430行で一致。
        残る1行は薬VAN側の丸め誤差0.52円）。列が無い・空のときだけ 在庫数×薬価 で補う。
        ※ 旧仕様の『過剰数金額』＝ 過剰数 × 薬価 とは別物です。 """
    a = g(row, '薬価金額')
    if str(a).strip() != '':
        return parse_num(a)
    return parse_num(g(row, '在庫数')) * parse_num(g(row, '薬価'))


def is_shortage(row):
    """ 受け手（不足候補）か：安全在庫数>0 かつ 在庫数<安全在庫数 """
    safe = parse_num(g(row, '安全在庫数'))
    zaiko = parse_num(g(row, '在庫数'))
    return safe > 0 and zaiko < safe


def holds_excess(row):
    """ 受け取り側が『すでに過剰保有』か（引取候補 tier②適正 と tier③参考 の分かれ目）。
        過剰在庫区分が非空 または 不動区分が非空 または 過剰数>0 なら True。
        ※ 期限切迫は「使えば消える」ため、ここでは過剰保有に含めない（出し手判定 is_supplier とは別定義）。 """
    for c in CONFIG['holds_excess_flag_cols']:
        if is_text_flag(g(row, c)):
            return True
    if parse_num(g(row, CONFIG['holds_excess_qty_col'])) > 0:
        return True
    return False


def is_legal_excluded(row):
    """ 麻薬・覚醒剤に該当するか（該当なら融通提案から除外） """
    for c in CONFIG['legal_exclude_cols']:
        if is_legal_flag(g(row, c)):
            return True
    return False


def warn_labels(row):
    """ 向精神薬・毒薬・劇薬など「要記録」警告の文字列を作る """
    out = []
    for c, label in CONFIG['legal_warn_cols'].items():
        if is_legal_flag(g(row, c)):
            out.append(label)
    return '・'.join(out)


def row_key(row):
    """ 突合キーを返す（個別医薬品CD優先、無ければレセプト電算CD、両方空なら None） """
    k = g(row, CONFIG['col_key'])
    if k:
        return k, CONFIG['col_key']
    k2 = g(row, CONFIG['col_key_fallback'])
    if k2:
        return k2, CONFIG['col_key_fallback']
    return None, None


def usage_qty6(row):
    """ 直近6ヶ月の出庫数合計。
    薬VANの出庫数は「払い出し＝マイナス」で記録されるため、この合計は通常マイナスになる。
    消化ペース（月あたり消化量）は、この値の絶対値を使って求める（build_candidates の pace 参照）。
    ※ 返品等でまれに正値が混じっても、使用中の判定は“出庫回数”で行うため実害は出ない。 """
    return sum(parse_num(g(row, c)) for c in CONFIG['outflow_qty_cols_6m'])


def usage_cnt6(row):
    """ 直近6ヶ月の出庫回数合計（払い出した回数＝プラス。使用中かどうかの判定に使う） """
    return sum(parse_num(g(row, c)) for c in CONFIG['outflow_cnt_cols_6m'])


def preprocess_keep(row):
    """ 前処理：削除/取扱/OTC で除外すべきなら False（=残さない） """
    if CONFIG['exclude_deleted'] and g(row, '削除フラグ').upper() == 'TRUE':
        return False
    if CONFIG['exclude_untreated'] and g(row, '取扱フラグ').upper() == 'FALSE':
        return False
    if CONFIG['exclude_otc'] and g(row, 'OTCフラグ').upper() == 'TRUE':
        return False
    return True


def slim_row(row):
    """ アプリ版でGシートへ入れる前に、マッチングに要る約35列（KEEP_COLS）だけへ圧縮する。
        列は列名で特定するので、元ファイルの列順は問わない。無い列は空文字で補う。 """
    return {c: g(row, c) for c in KEEP_COLS}


def slim_rows(rows):
    """ 行のリストをまとめて slim_row する。 """
    return [slim_row(r) for r in rows]


# ============================================================================
# ファイル名の解釈
# ============================================================================
def parse_filename(path_or_name):
    """ 「店舗名_YYYYMM.csv」から 店舗名 と YYYYMM を取り出す（パスでもファイル名でも可）。 """
    base = os.path.splitext(os.path.basename(path_or_name))[0]
    if '_' in base:
        name, tail = base.rsplit('_', 1)
        if len(tail) == 6 and tail.isdigit():
            return name, tail
        # 末尾がYYYYMMでない場合は全体を店名扱い
        return base, None
    return base, None


# ============================================================================
# ファイル読み込み（「バイト＋ファイル名」を入口に一般化）
#   ・コマンド版：パスからバイトを読み、read_store_header / read_store_rows を呼ぶ
#   ・アプリ版　：アップロード物のバイトをそのまま read_*_from_bytes へ渡す
#   どちらも下の共通ロジックに合流するので、計算は完全に一致する。
# ============================================================================
def _norm_number(v):
    """ 数値を、CSV由来の値と揃う文字列にする。
        整数はそのまま（例 202.0→'202'）、端数があれば余計な末尾ゼロを落とす（例 14649.60→'14649.6'）。
        後段の parse_num が float 化するので、多少の体裁差は結果に影響しない。 """
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v).strip()
    if f != f:  # NaN
        return ''
    if f.is_integer():
        return str(int(f))
    # 10桁で丸めて浮動小数点の“ゴミ”を消してから末尾ゼロ・小数点を除去
    return ('%.10f' % f).rstrip('0').rstrip('.')


def _norm_xldate(v, datemode):
    """ Excelのシリアル値（例 46419.0）を 'YYYY/M/D'（ゼロ埋めなし＝CSVと同じ体裁）に変換する。
        有効期限・最終出庫日などの日付列はこれで既存処理（parse_date／そのまま表示）に載る。 """
    import xlrd
    dt = xlrd.xldate.xldate_as_datetime(v, datemode)
    return '%d/%d/%d' % (dt.year, dt.month, dt.day)


def _xls_cell_to_str(cell, datemode):
    """ xlrd のセル1つを、CSV DictReader と同じ「文字列」に正規化する。
        ・空欄／エラー → ''
        ・文字列 → そのまま
        ・数値 → 文字列化（_norm_number）
        ・日付（シリアル値）→ 'YYYY/M/D'（_norm_xldate）
        ・真偽（削除／取扱／OTCフラグ等）→ CSVに合わせて 'TRUE' / 'FALSE' """
    import xlrd
    t = cell.ctype
    v = cell.value
    if t == xlrd.XL_CELL_TEXT:            # 1: 文字列
        return str(v).strip()
    if t == xlrd.XL_CELL_NUMBER:          # 2: 数値
        return _norm_number(v)
    if t == xlrd.XL_CELL_DATE:            # 3: 日付（Excelシリアル値）
        return _norm_xldate(v, datemode)
    if t == xlrd.XL_CELL_BOOLEAN:         # 4: 真偽（薬VANのフラグ列はここ）
        return 'TRUE' if v else 'FALSE'
    # 0:空欄, 5:エラー, 6:ブランク など
    return ''


def _norm_openpyxl_value(v):
    """ openpyxl（.xlsx）のセル値を、CSV DictReader と同じ「文字列」に正規化する。 """
    import datetime as _dt
    if v is None:
        return ''
    if isinstance(v, bool):               # ※bool は int より先に判定する
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float)):
        return _norm_number(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return '%d/%d/%d' % (v.year, v.month, v.day)
    return str(v).strip()


# --- .csv（バイト → 文字コードを自動判定して復号。newline='' で開いていた挙動に合わせる）---
# 薬VANのCSVは店（出力の設定）によって UTF-8(BOM付き) と Shift-JIS(cp932) の2種類があるため、
# UTF-8 で読めなければ cp932 で読み直す。UTF-8 は不正なバイト列を確実に弾くので、この順なら誤判定しない。
_CSV_ENCODINGS = ('utf-8-sig', 'cp932')


def _decode_csv_bytes(data):
    """ CSVのバイト列を文字列に復号する。UTF-8（BOM有無どちらも）→ Shift-JIS(cp932) の順に試す。 """
    last_err = None
    for enc in _CSV_ENCODINGS:
        try:
            return data.decode(enc)
        except UnicodeDecodeError as e:
            last_err = e
    raise ValueError(
        'CSVの文字コードを判別できませんでした（UTF-8 でも Shift-JIS でもありません）。'
        '詳細：%s' % last_err
    )


def _read_csv_header_bytes(data):
    text = _decode_csv_bytes(data)
    return list(next(csv.reader(io.StringIO(text, newline=''))))


def _read_csv_rows_bytes(data):
    text = _decode_csv_bytes(data)
    return list(csv.DictReader(io.StringIO(text, newline='')))


# --- .xls（xlrd。file_contents= でバイトから直接開く。シートは "QFORM" 優先、無ければ先頭）---
def _open_xls_sheet_bytes(data):
    import xlrd
    book = xlrd.open_workbook(file_contents=data)
    names = book.sheet_names()
    sheet = book.sheet_by_name('QFORM') if 'QFORM' in names else book.sheet_by_index(0)
    return book, sheet


def _read_xls_header_bytes(data):
    book, sh = _open_xls_sheet_bytes(data)
    if sh.nrows < 1:
        return []
    return [str(sh.cell_value(0, c)) for c in range(sh.ncols)]


def _read_xls_rows_bytes(data):
    book, sh = _open_xls_sheet_bytes(data)
    if sh.nrows < 1:
        return []
    headers = [str(sh.cell_value(0, c)) for c in range(sh.ncols)]
    rows = []
    for rr in range(1, sh.nrows):
        d = {}
        for ci, h in enumerate(headers):
            d[h] = _xls_cell_to_str(sh.cell(rr, ci), book.datemode)
        rows.append(d)
    return rows


# --- .xlsx（openpyxl。BytesIO でバイトから直接開く。シートは "QFORM" 優先、無ければ先頭）---
def _open_xlsx_sheet_bytes(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb['QFORM'] if 'QFORM' in wb.sheetnames else wb[wb.sheetnames[0]]
    return wb, ws


def _read_xlsx_header_bytes(data):
    wb, ws = _open_xlsx_sheet_bytes(data)
    try:
        first = next(ws.iter_rows(values_only=True))
    except StopIteration:
        wb.close()
        return []
    header = [('' if h is None else str(h)) for h in first]
    wb.close()
    return header


def _read_xlsx_rows_bytes(data):
    wb, ws = _open_xlsx_sheet_bytes(data)
    it = ws.iter_rows(values_only=True)
    try:
        first = next(it)
    except StopIteration:
        wb.close()
        return []
    headers = [('' if h is None else str(h)) for h in first]
    rows = []
    for raw in it:
        d = {}
        for ci, h in enumerate(headers):
            v = raw[ci] if ci < len(raw) else None
            d[h] = _norm_openpyxl_value(v)
        rows.append(d)
    wb.close()
    return rows


def _ext_of(name):
    return os.path.splitext(name)[1].lower()


def read_header_from_bytes(filename, data):
    """ 入力ファイルの見出し行（列名のリスト）を、形式に依らず返す（バイト入力）。 """
    ext = _ext_of(filename)
    if ext == '.csv':
        return _read_csv_header_bytes(data)
    if ext == '.xls':
        return _read_xls_header_bytes(data)
    if ext == '.xlsx':
        return _read_xlsx_header_bytes(data)
    raise ValueError('対応していない拡張子です（.csv / .xls / .xlsx のみ対応）：%s' % ext)


def read_rows_from_bytes(filename, data):
    """ 入力ファイルの全データ行を、形式に依らず「列名→文字列」の辞書のリストで返す（バイト入力）。 """
    ext = _ext_of(filename)
    if ext == '.csv':
        return _read_csv_rows_bytes(data)
    if ext == '.xls':
        return _read_xls_rows_bytes(data)
    if ext == '.xlsx':
        return _read_xlsx_rows_bytes(data)
    raise ValueError('対応していない拡張子です（.csv / .xls / .xlsx のみ対応）：%s' % ext)


# --- パス入口（コマンド版はここを呼ぶ。中身は「バイトを読んで上の関数へ」）---
def read_store_header(path):
    """ 入力ファイルの見出し行（列名のリスト）を、形式に依らず返す（パス入力）。 """
    with open(path, 'rb') as f:
        data = f.read()
    return read_header_from_bytes(os.path.basename(path), data)


def read_store_rows(path):
    """ 入力ファイルの全データ行を、形式に依らず辞書のリストで返す（パス入力）。 """
    with open(path, 'rb') as f:
        data = f.read()
    return read_rows_from_bytes(os.path.basename(path), data)


# ============================================================================
# 引取候補店の文字列を作る
# ============================================================================
def _join_with_overflow(labels):
    """ 候補ラベルを最大 max_candidates 件まで ' / ' 連結し、あふれは「他N店」でまとめる。 """
    if not labels:
        return ''
    cap = CONFIG['max_candidates']
    shown = labels[:cap]
    if len(labels) > cap:
        shown.append('他%d店' % (len(labels) - cap))
    return ' / '.join(shown)


def build_candidates(by_key, key, source_store, supply_qty, exp_date, base_date):
    """ 出し手（デッド／期限切迫）1件について、引取候補店を①②③のtierに分けて
        (本命候補の文字列, 参考(過剰)候補の文字列) の2つを返す。
          本命 ＝ ①不足中 → ②使用中(適正) の順（最大 max_candidates 件、あふれは「他N店」）
          参考 ＝ ③使用中(過剰)。本命とは別の列に出すことで、本命候補と混ざらないようにする。 """
    if not key:
        return '（医薬品コード無し・突合対象外）', ''
    entries = by_key.get(key, [])
    short = []    # ① 不足中
    use_ok = []   # ② 使用中(適正)：使っていて、かつ受け取り側が過剰保有していない＝本命
    use_ref = []  # ③ 使用中(過剰)：使っているが受け取り側も過剰保有＝参考（下位・別枠）
    for e in entries:
        if e['store'] == source_store:
            continue
        if e['shortage']:
            short.append(e)
        elif e['usage_cnt6'] > 0:
            # 使用中：直近6ヶ月に出庫回数がある（＝その薬を実際に使っている店）。
            #   判定は符号に左右されない“出庫回数合計>0”で行う。
            #   さらに、受け取り側が既に過剰保有かどうかで「適正(本命)」と「過剰(参考)」に分ける。
            if e['holds_excess']:
                use_ref.append(e)
            else:
                use_ok.append(e)

    def pace(e):
        # 月あたり消化量。薬VANの出庫数はマイナス＝出庫なので、絶対値をとってから6で割る。
        return abs(e['usage_qty6']) / 6.0

    def detail_for(e):
        p = pace(e)
        if p <= 0:
            # 出庫回数はあるが出庫数の正味がゼロ（返品と相殺等）＝消化量が読めないケース
            return '消化ペース不明'
        # 消化目安＝渡す数量（在庫数）÷ 相手の月あたり消化量
        months = supply_qty / p
        if exp_date and base_date:
            remain = month_diff(base_date, exp_date)
            if months <= remain:
                return '約%sヶ月で消化可' % fmt_months(months)
            return '約%sヶ月・期限内消化不可' % fmt_months(months)
        return '約%sヶ月' % fmt_months(months)

    short.sort(key=lambda e: -pace(e))
    use_ok.sort(key=lambda e: -pace(e))
    use_ref.sort(key=lambda e: -pace(e))

    L = CONFIG['tier_labels']
    # 本命（①不足中 → ②使用中(適正)）
    main_ordered = [(L['short'], e) for e in short] + [(L['use_ok'], e) for e in use_ok]
    main_labels = ['%s（%s・%s）' % (e['store'], tag, detail_for(e)) for tag, e in main_ordered]
    main_str = _join_with_overflow(main_labels) or '該当なし'

    # 参考（③使用中(過剰)）… 必ず末尾・別枠。ラベルで本命と一目で区別できるようにする。
    ref_labels = ['%s（%s・%s）' % (e['store'], L['use_ref'], detail_for(e)) for e in use_ref]
    ref_str = _join_with_overflow(ref_labels)

    return main_str, ref_str


# ============================================================================
# 予約（受け手の店が「この品はうちが引き取ります」と押さえた印）の小道具
#   ・予約1件は (出し手店, 予約キー) で1つ。予約キーは除外キー（exclusion_key）と同じ。
#   ・「品目まるごと」を押さえる方式（本間部長判断 2026-07-28）。数量の指定はしない
#     ＝何錠もらうかの相談は従来どおり電話・デスクネッツで行う。
# ============================================================================
def reservation_key(source_store, row):
    """ 予約1件を特定するキー (出し手店, 品目キー) を返す。★除外と同じ品目キーを使う。 """
    return (source_store, exclusion_key(row))


def _reserved_by(reserved, source_store, ex_key):
    """ その品を予約している店の名前を返す（予約が無ければ空文字）。 """
    if not reserved:
        return ''
    r = reserved.get((source_store, ex_key))
    return (r or {}).get('店', '') or ''


def _reserve_label(reserved, source_store, ex_key):
    """ Excel／Gシートの『予約』列に入れる文字列（例：'東立石 2026/07/28 11:40（受取：3ヶ月後）'）。
        ★2026-08-01：受取時期（今すぐ／Nヶ月後）を末尾に足した。予約を翌月以降へ持ち越せる
          ようになったので、Gシート・Excel の記録だけを見ても「いつ引き取るか」が分かるようにする。 """
    if not reserved:
        return ''
    r = reserved.get((source_store, ex_key))
    if not r:
        return ''
    base = ('%s %s' % (r.get('店', ''), r.get('日時', ''))).strip()
    pickup = r.get('受取ラベル', '')
    if pickup:
        base = '%s（受取：%s）' % (base, pickup)
    return base


def _reserved_pickup(reserved, source_store, ex_key):
    """ その品の予約の『受取ラベル』（例 '3ヶ月後'／'今すぐ'）を返す（予約が無ければ空文字）。
        受取ラベルは reservation_map（当月を知っている app_logic 側）で作って渡してもらう
        ＝当月からの月数を計算できる場所を1つに保ち、ここ（yuzu_core）は受け取るだけにする。 """
    if not reserved:
        return ''
    r = reserved.get((source_store, ex_key))
    return (r or {}).get('受取ラベル', '') or ''


def build_candidate_names(key, entries):
    """ 引取候補店を「店名だけ」つないだ文字列を返す（②③の画面表示用。2026-07-28 本間部長指示）。

        画面の『引取候補店』列は店名だけを出す。理由：店がやることは「その店に電話する」で、
        tier（不足中／使用中）や消化目安まで並ぶと横に長くなり、肝心の店名が読み取りづらい。
        ★Excel／Gシートの『引取候補店』列は従来どおり tier・消化目安つきのまま（本部が優先順位を見るため）。

        ★build_candidates が作った文字列から店名を切り出してはいけない。
          あの文字列は上位 max_candidates 店で「他N店」に丸められるため、解析すると候補が漏れる。
          そこで丸め前の候補リスト（build_candidate_entries の戻り値）から組み立てる。
          並び順（①不足中→②使用中、各グループ内は消化ペースの速い順）と丸めの仕方は
          build_candidates とまったく同じなので、店名の並びは詳細版と一致する。

        文言も build_candidates とそろえる：
          医薬品コード無し → '（医薬品コード無し・突合対象外）' ／ 候補ゼロ → '該当なし' """
    if not key:
        return '（医薬品コード無し・突合対象外）'
    return _join_with_overflow([e['引取候補店'] for e in entries]) or '該当なし'


def build_candidate_entries(by_key, key, source_store, supply_qty, exp_date, base_date):
    """ 出し手（デッド／期限切迫）1件について、引取候補店を「候補ごとに1件の辞書」で返す。
        ④受け手ビュー（他店が出そうとしていて、自店が引き取れば活かせる品の一覧）の土台です。

        ★build_candidates（表示用の文字列を作る関数）と選別ロジックは同じですが、
          この関数は _join_with_overflow を通しません＝上位5店（CONFIG['max_candidates']）で
          丸めず、拾った候補を丸め無しで全件返します。
          （build_candidates の文字列は6店目以降を「他N店」に丸めるため、その文字列を
            解析して④を作ると候補漏れが起き、しかも画面には「該当なし」と出て漏れに気づけません。
            そのため、丸めを通さない構造化データをここで別に作ります。）

        拾うのは ①不足中（short）と ②使用中(適正)（use_ok）の2つだけです。
          ③参考（過剰だが使用中／use_ref）は含めません
          （本間部長確定：④の「なぜ候補か」は 不足中／使用中 の2値）。

        返す辞書は「候補ごとに違う情報」だけに絞ります（薬品名・出し手の在庫数などの品目情報は持たせません）。
          理由：同じ値を2箇所で作ると、将来どちらかを直したときに静かにズレます。
                品目情報は呼び出し側（compute_matching）が提案行から借りて付け足します（値の出どころを1つに保つ）。

        ★2026-08-01 追加：候補ごとに違う情報として『自店の在庫数』を持たせます（④の新しい列）。
          値は候補店 e の e['over_qty'] を借ります。over_qty は store_key_info で組んだ
          「その(店,キー)の在庫数（＝stock_qty）」で、同じ(店,キー)が複数行（ロット別様式）でも
          合算済みです。★ここで e['zaiko'] を使ってはいけません。zaiko は store_key_info の
          else 分岐（同じ(店,キー)の2行目以降）で合算されず捨てられており、ロット別で複数行に
          分かれている店では過少になります。over_qty は合算されているので正しい在庫数になります。

        戻り値は辞書のリスト（候補が無ければ空リスト）：
          {'引取候補店': 店名, 'なぜ候補か': '不足中' or '使用中',
           '_tier_order': 0(不足中) or 1(使用中), '消化目安': 目安文字列,
           '自店の在庫数': 候補店(=④では自店)がいまその品を持っている在庫数}
        ※ key が空（医薬品コード無し・突合対象外）のときは空リストを返します
          （build_candidates が '（医薬品コード無し・突合対象外）' を返すのと同じケース）。 """
    if not key:
        return []
    entries = by_key.get(key, [])
    short = []    # ① 不足中
    use_ok = []   # ② 使用中(適正)：使っていて、かつ受け取り側が過剰保有していない＝本命
    for e in entries:
        if e['store'] == source_store:
            # 自店は引取候補にしない（build_candidates と同じガード）
            continue
        if e['shortage']:
            short.append(e)
        elif e['usage_cnt6'] > 0:
            # 使用中：直近6ヶ月に出庫回数がある店。受け取り側が過剰保有していない品だけを②使用中に入れる。
            #   （過剰保有している品＝③参考は build_candidates 側で扱う。ここでは拾わない）
            if not e['holds_excess']:
                use_ok.append(e)

    def pace(e):
        # 月あたり消化量。薬VANの出庫数はマイナス＝出庫なので、絶対値をとってから6で割る。
        return abs(e['usage_qty6']) / 6.0

    def detail_for(e):
        # ★build_candidates 内の detail_for と一字一句同じ計算式（消化目安の二重管理を避ける）。
        p = pace(e)
        if p <= 0:
            # 出庫回数はあるが出庫数の正味がゼロ（返品と相殺等）＝消化量が読めないケース
            return '消化ペース不明'
        # 消化目安＝渡す数量（在庫数）÷ 相手の月あたり消化量
        months = supply_qty / p
        if exp_date and base_date:
            remain = month_diff(base_date, exp_date)
            if months <= remain:
                return '約%sヶ月で消化可' % fmt_months(months)
            return '約%sヶ月・期限内消化不可' % fmt_months(months)
        return '約%sヶ月' % fmt_months(months)

    short.sort(key=lambda e: -pace(e))
    use_ok.sort(key=lambda e: -pace(e))

    L = CONFIG['tier_labels']
    out = []
    # ①不足中（_tier_order=0）を先に、②使用中（_tier_order=1）を後に並べる（既存の本命順と同じ）
    #   ★『自店の在庫数』＝候補店 e の over_qty（合算済みの在庫数）。zaiko は使わない（上の docstring 参照）。
    for e in short:
        out.append({'引取候補店': e['store'], 'なぜ候補か': L['short'],
                    '_tier_order': 0, '消化目安': detail_for(e), '自店の在庫数': e['over_qty']})
    for e in use_ok:
        out.append({'引取候補店': e['store'], 'なぜ候補か': L['use_ok'],
                    '_tier_order': 1, '消化目安': detail_for(e), '自店の在庫数': e['over_qty']})
    return out


def build_candidate_entries_ref(by_key, key, source_store, supply_qty, exp_date, base_date):
    """ ③参考（過剰だが使用中／use_ref）を「候補ごとに1件の辞書」で丸め無しで返す（2026-08-01 追加）。
        ④の別枠『いまは在庫があるが、先になら引き取れる薬』の土台です。

        build_candidate_entries（①不足中②使用中）と対になる関数で、こちらは use_ref だけを拾います。
          use_ref ＝ その薬を使ってはいる（6ヶ月出庫回数>0）が、受け取り側も過剰保有している店。
          今すぐ送ると移した先で新しいデッドを作りかねないため、本来の④（今すぐ引き取れる薬）からは
          意図的に除外されている（2026-07-27 本間部長確定）。この判断は覆さず、別枠に分けて
          「今は在庫があるが、いまの在庫を使い切る先の時期なら引き取れる」品として出します。

        ★build_candidate_entries と同じく _join_with_overflow（丸め）を通しません。
          上位5店で『他N店』に丸めた文字列から作ると候補が漏れ、しかも画面に『該当なし』と出て
          漏れに気づけないため、必ず丸め前の構造化データから作ります。

        戻り値は辞書のリスト（候補が無ければ空リスト）：
          {'引取候補店': 店名, 'なぜ候補か': '参考:過剰だが使用中',
           '_tier_order': 2, '消化目安': 目安文字列, '自店の在庫数': 候補店の在庫数}
        ※ 自店の在庫数の取り方（over_qty を借りる・zaiko は使わない）は build_candidate_entries と同じ。 """
    if not key:
        return []
    entries = by_key.get(key, [])
    use_ref = []   # ③ 使用中(過剰)：使っているが受け取り側も過剰保有＝参考
    for e in entries:
        if e['store'] == source_store:
            # 自店は引取候補にしない（build_candidates と同じガード）
            continue
        if e['shortage']:
            # 不足中は本来の④（build_candidate_entries）の担当。ここでは拾わない
            continue
        if e['usage_cnt6'] > 0 and e['holds_excess']:
            use_ref.append(e)

    def pace(e):
        # 月あたり消化量。薬VANの出庫数はマイナス＝出庫なので、絶対値をとってから6で割る。
        return abs(e['usage_qty6']) / 6.0

    def detail_for(e):
        # ★build_candidates / build_candidate_entries 内の detail_for と一字一句同じ計算式
        #   （既存2関数がすでに同じ式を持ち「必ず同じにする」約束のため、それに合わせる）。
        p = pace(e)
        if p <= 0:
            return '消化ペース不明'
        months = supply_qty / p
        if exp_date and base_date:
            remain = month_diff(base_date, exp_date)
            if months <= remain:
                return '約%sヶ月で消化可' % fmt_months(months)
            return '約%sヶ月・期限内消化不可' % fmt_months(months)
        return '約%sヶ月' % fmt_months(months)

    use_ref.sort(key=lambda e: -pace(e))
    L = CONFIG['tier_labels']
    out = []
    for e in use_ref:
        out.append({'引取候補店': e['store'], 'なぜ候補か': L['use_ref'],
                    '_tier_order': 2, '消化目安': detail_for(e), '自店の在庫数': e['over_qty']})
    return out


# ============================================================================
# 全店一括計算 compute_matching(stores)
#   入力 stores = [{'name': 店名, 'ym': 'YYYYMM'|None, 'base_date': date, 'rows': [dict,...]}, ...]
#        （rows は前処理 preprocess_keep 済みの行。列名→文字列の辞書。）
#   戻り値 dict … 4シート相当データ＋自己検算＋ログ用の集計。
#   ※ 現在 create_yuzu_list.main() にベタ書きされていた計算を、そのまま関数化したものです。
#     計算内容は一切変えていません（回帰で1円も変わらないことを保証する目的）。
# ============================================================================
def compute_matching(stores, excluded=None, reserved=None):
    """ excluded … 店が「融通に出さない」と外した品目の集合。
        {(店名, exclusion_key(row)), ...} の形。ここで出し手から完全に取り除くので、
        自店の②③だけでなく全店一覧・マトリクス・受け手の供給元・Excel・サマリ・
        自己検算のすべてから同時に消える（どこかに残って他店から問い合わせが来る事故を防ぐ）。

        reserved … 受け手の店が「この品はうちが引き取ります」と予約した品。
        {(出し手店, exclusion_key(row)): {'店': 予約した店, '日時': 'YYYY/MM/DD HH:MM',
                                          '受取予定月': 'YYYYMM', '受取ラベル': '3ヶ月後'|'今すぐ'}, ...} の形。
          ★受取予定月・受取ラベルは2026-08-01追加（予約を最大3ヶ月先まで持ち越せるように）。
            当月からの月数を計算できる app_logic.reservation_map（当月を知っている場所）で作って渡す。
        ★除外と違い、ここでは出し手から取り除かない（＝件数・金額・自己検算は1円も動かない）。
          予約は「もう相手が決まった」という印であって、品が消えるわけではないため。
            ・出し手の②③  … 引取候補店の欄が「◯◯が引取予定（受取：3ヶ月後）」に変わる（誰にいつ渡すか分かる）
            ・受け手の④    … 予約した店だけに残り、ほかの店の④からは消える
            ・Excel／Gシート… 融通提案シートの『予約』列に「店名 日時（受取：◯）」が入る
          そのため予約キーは除外キー（exclusion_key）と同じものを使う＝品目の呼び名を1つに保つ。 """
    excluded = set(excluded or ())
    reserved = dict(reserved or {})

    def is_excluded(store_name, row):
        return bool(excluded) and (store_name, exclusion_key(row)) in excluded

    # --- 区分値の分布（ログ用・全店合算）---
    dist = {}
    for col in DIST_COLS:
        cnt = Counter()
        for s in stores:
            for row in s['rows']:
                cnt[g(row, col) or '（空欄）'] += 1
        dist[col] = cnt

    # --- (店,キー)ごとに集約 → by_key を作る ---
    store_key_info = {}
    nokey_rows = 0
    for s in stores:
        for row in s['rows']:
            key, _ = row_key(row)
            if key is None:
                nokey_rows += 1
                continue
            kk = (s['name'], key)
            # 出し手から外す2条件：店が「出さない」と外した品／少額（1,500円未満）の品
            ex = is_excluded(s['name'], row) or is_below_min_amount(row)
            sh = is_shortage(row)
            dead = is_dead(row) and not ex     # デッド（不動区分が非空）
            expf = is_expiry(row, s['base_date']) and not ex  # 期限切迫（有効期限まで5ヶ月以内）
            supp = dead or expf                # 出し手（デッド or 期限切迫）
            he = holds_excess(row)   # 受け取り側の「過剰保有」（tier②適正/③参考の分かれ目）
            uq = usage_qty6(row)
            uc = usage_cnt6(row)
            over_qty = stock_qty(row)      # 在庫数（旧：過剰数）
            over_amt = stock_amount(row)   # 在庫金額＝在庫数×薬価（旧：過剰数金額）
            exp = parse_date(g(row, '有効期限'))
            name = g(row, '薬品名')
            zaiko = parse_num(g(row, '在庫数'))
            safe = parse_num(g(row, '安全在庫数'))
            e = store_key_info.get(kk)
            if e is None:
                store_key_info[kk] = {
                    'store': s['name'], 'shortage': sh,
                    'dead': dead, 'expiry': expf, 'supplier': supp,
                    'holds_excess': he,
                    'usage_qty6': uq, 'usage_cnt6': uc, 'over_qty': over_qty,
                    'over_amt': over_amt, 'exp': exp, 'name': name,
                    'zaiko': zaiko, 'safe': safe}
            else:
                e['shortage'] = e['shortage'] or sh
                e['dead'] = e['dead'] or dead
                e['expiry'] = e['expiry'] or expf
                e['supplier'] = e['supplier'] or supp
                e['holds_excess'] = e['holds_excess'] or he
                e['usage_qty6'] += uq
                e['usage_cnt6'] += uc
                e['over_qty'] += over_qty
                e['over_amt'] += over_amt
                if exp and (e['exp'] is None or exp < e['exp']):
                    e['exp'] = exp

    by_key = defaultdict(list)
    for (store, key), e in store_key_info.items():
        by_key[key].append(e)

    # --- 出し手（融通提案）を1行ずつ作る ---
    #   出し手＝デッド または 期限切迫（A案）。純粋な過剰は提案に載せない。
    #   各行に「種別」（デッド／期限切迫・排他デッド優先）を持たせる。
    proposal_rows = []
    # ④受け手ビューの土台：引取候補店を「候補ごとに1件」で貯めるリスト（丸め無し・下のループで足す）
    candidate_rows = []
    # ④の別枠『いまは在庫があるが、先になら引き取れる薬』の土台（use_ref＝過剰だが使用中）。
    #   ★candidate_rows（本来の④）とは完全に別のリストにして、既存④に出る品目は1件も変えない。
    #   ★このリストは Gシート／Excel には書き戻さない（結果4タブの payload に含めない）。
    candidate_rows_ref = []
    legal_excluded_count = 0
    legal_excluded_amt = 0.0
    nokey_overstock = 0
    user_excluded_count = 0
    small_excluded_count = 0
    small_excluded_amt = 0.0
    # 少額カットの店別内訳（{店名: {'count': 件数, 'amt': 金額}}）。
    #   画面（②③）で「自店の少額非表示は何件・いくらか」を出すために足す。
    #   全店合計の small_excluded_count / small_excluded_amt は従来どおり別に持つ（変更しない）。
    small_by_store = {}
    for s in stores:
        for row in s['rows']:
            if not is_supplier(row, s['base_date']):
                continue
            if is_excluded(s['name'], row):
                # 店が「この品は出さない」と外したもの
                user_excluded_count += 1
                continue
            if is_below_min_amount(row):
                # 少額（既定1,500円未満）で載せない品
                small_excluded_count += 1
                small_excluded_amt += stock_amount(row)
                # 店別内訳にも同じ品を足す（全店合計は上の2変数のまま・二重管理しない）
                sb = small_by_store.setdefault(s['name'], {'count': 0, 'amt': 0.0})
                sb['count'] += 1
                sb['amt'] += stock_amount(row)
                continue
            if is_legal_excluded(row):
                legal_excluded_count += 1
                legal_excluded_amt += stock_amount(row)
                continue
            key, _ = row_key(row)
            if key is None:
                nokey_overstock += 1
            over_qty = stock_qty(row)      # 在庫数
            over_amt = stock_amount(row)   # 在庫金額
            exp = parse_date(g(row, '有効期限'))
            cand_main, cand_ref = build_candidates(
                by_key, key, s['name'], over_qty, exp, s['base_date'])
            # 丸め無しの候補リスト。④受け手ビューの土台と、②③画面用の「店名だけ」の文字列の
            #   両方をここから作る（同じ候補を2度計算しない／出どころを1つに保つ）。
            cand_entries = build_candidate_entries(
                by_key, key, s['name'], over_qty, exp, s['base_date'])
            # ④の別枠『先になら引き取れる薬』用の候補（use_ref＝過剰だが使用中）。丸め無し。
            #   本来の④（cand_entries）とは別のリストへ入れるので、既存④の品目には一切影響しない。
            cand_entries_ref = build_candidate_entries_ref(
                by_key, key, s['name'], over_qty, exp, s['base_date'])
            remain = month_diff(s['base_date'], exp) if exp else None
            pr = {
                '出し手店': s['name'], '種別': supplier_category(row, s['base_date']),
                '薬品名': g(row, '薬品名'), '単位': g(row, '単位'),
                'メーカ名': g(row, 'メーカ名'), '在庫数': round(over_qty, 2),
                '在庫金額': round(over_amt, 2),
                '過剰在庫区分': g(row, '過剰在庫区分'), '不動区分': g(row, '不動区分'),
                '期限切迫区分': g(row, '期限切迫区分'), '有効期限': fmt_date(exp),
                'ロットNO': g(row, 'ロットNO'), '最終出庫日': g(row, '最終出庫日'),
                '区分': warn_labels(row),
                '引取候補店': cand_main, '参考:過剰だが使用中の店': cand_ref,
                # 画面（②③）用の店名だけの版。Excel／Gシートは上の『引取候補店』（詳細つき）を使う。
                '引取候補店（店名のみ）': build_candidate_names(key, cand_entries),
                '6ヶ月出庫回数': round(usage_cnt6(row), 2), '医薬品CD': key or '',
                '_ex_key': exclusion_key(row),
                # 予約（受け手の店が「うちが引き取ります」と押さえた印）。
                #   '予約'   … Excel／Gシートに出す文字列（例：'東立石 2026/07/28 11:40（受取：3ヶ月後）'）
                #   '_予約店' … 画面のしぼり込みに使う店名だけ（予約が無ければ空文字）
                #   '_予約受取' … 受取時期のラベル（例 '3ヶ月後'／'今すぐ'）。②③画面の「◯◯が引取予定」に足す。
                '予約': _reserve_label(reserved, s['name'], exclusion_key(row)),
                '_予約店': _reserved_by(reserved, s['name'], exclusion_key(row)),
                '_予約受取': _reserved_pickup(reserved, s['name'], exclusion_key(row)),
                # 期限切迫フラグ（＝赤ハイライトの元）。有効期限が読めて、基準日から5ヶ月以内なら True。
                #   ★is_expiry と同じ計算（remain＝基準日→有効期限の月数）に一本化しているので、
                #     「載せる条件」と「赤の条件」がズレない（A案・2026-08-04）。
                '_expiry_flag': (remain is not None
                                 and remain <= CONFIG['expiry_within_months']),
                # 滞留（何ヶ月つづけて載っているか）。ここでは「今月から」を初期値として置き、
                #   過去の記録がある場合だけ app_logic.apply_stagnation が上書きする。
                #   ★初期値を必ず入れておくのは、履歴が読めなかった月でも画面・Excelが
                #     欠損キーで落ちないようにするため（履歴は「あれば良くなる」情報に留める）。
                '滞留': '', '_滞留月数': 1, '_滞留区分': 'new', '_先月予約': False,
                # 引取候補店がいるか。★文字列（'該当なし' 等）を見て判定すると
                #   文言を変えた瞬間に壊れるので、候補リストの中身そのもので持たせる。
                '_候補あり': bool(cand_entries),
                '_remain': remain, '_amt_raw': over_amt}
            proposal_rows.append(pr)
            # ④受け手ビューの土台：この出し手行の引取候補店を「候補ごとに1件」で作り（丸め無し）、
            #   品目情報（薬品名・在庫数など）は上で作った提案行 pr から借りてマージして貯める。
            #   ★ここは少額カット（1,500円未満）・除外・法規制除外を全部くぐり抜けた行だけが到達するので、
            #     それらの除外は候補にも自動で反映される（この地点に置いているのがミソ）。
            for ce in cand_entries:
                candidate_rows.append({
                    '引取候補店': ce['引取候補店'], 'なぜ候補か': ce['なぜ候補か'],
                    '消化目安':   ce['消化目安'],   '_tier_order': ce['_tier_order'],
                    '出し手店': pr['出し手店'], '薬品名': pr['薬品名'], '単位': pr['単位'],
                    '在庫数':   pr['在庫数'],   '在庫金額': pr['在庫金額'],
                    # ★『自店の在庫数』＝候補店（④では自店）がいまその品を持っている在庫数。
                    #   候補ごとに違う情報なので候補 ce から借りる（品目情報は上の pr から借りる）。
                    '自店の在庫数': round(ce['自店の在庫数'], 2),
                    '有効期限': pr['有効期限'], '区分': pr['区分'], '医薬品CD': pr['医薬品CD'],
                    '_remain':  remain,
                    # 滞留は出し手の提案行と同じ値を持たせる（④受け手の画面にも出すため）。
                    #   ★apply_stagnation は提案行を直したあとに候補行へ配り直すので、
                    #     ここでの値は初期値。実値の詰め直しは app_logic 側で行う。
                    '滞留': pr['滞留'], '_滞留月数': pr['_滞留月数'],
                    '_滞留区分': pr['_滞留区分'], '_先月予約': pr['_先月予約'],
                    # 予約済みなら、その店の④にだけ残して他店の④からは消すために持たせる
                    '_予約店': pr['_予約店'],
                    # ★予約を保存するときのキー。(出し手店, _ex_key) で1件を特定する。
                    #   ここで渡しておかないと、④から予約したときにキーが空のまま保存され、
                    #   読み込み時に捨てられて「予約したのに消える」ことになる。
                    '_ex_key': pr['_ex_key'],
                })
            # ④の別枠『先になら引き取れる薬』の土台。作りは candidate_rows とそっくりだが、
            #   拾う候補が use_ref（過剰だが使用中）だけ＝別リスト candidate_rows_ref に貯める。
            #   ★これにより既存④（candidate_rows）の中身はバイト単位で不変のまま、別枠を足せる。
            for ce in cand_entries_ref:
                candidate_rows_ref.append({
                    '引取候補店': ce['引取候補店'], 'なぜ候補か': ce['なぜ候補か'],
                    '消化目安':   ce['消化目安'],   '_tier_order': ce['_tier_order'],
                    '出し手店': pr['出し手店'], '薬品名': pr['薬品名'], '単位': pr['単位'],
                    '在庫数':   pr['在庫数'],   '在庫金額': pr['在庫金額'],
                    '自店の在庫数': round(ce['自店の在庫数'], 2),
                    '有効期限': pr['有効期限'], '区分': pr['区分'], '医薬品CD': pr['医薬品CD'],
                    '_remain':  remain,
                    '滞留': pr['滞留'], '_滞留月数': pr['_滞留月数'],
                    '_滞留区分': pr['_滞留区分'], '_先月予約': pr['_先月予約'],
                    '_予約店': pr['_予約店'],
                    '_ex_key': pr['_ex_key'],
                })
    proposal_rows.sort(key=lambda r: -r['在庫金額'])

    # --- 受け手（不足品目一覧）を作る ---
    shortage_rows = []
    for s in stores:
        for row in s['rows']:
            if not is_shortage(row):
                continue
            key, _ = row_key(row)
            safe = parse_num(g(row, '安全在庫数'))
            zaiko = parse_num(g(row, '在庫数'))
            # 供給元は「デッドまたは期限切迫で持つ他店」に統一（A案。旧・過剰保有基準は廃止）
            holders = []
            if key:
                for e in by_key.get(key, []):
                    if e['store'] == s['name']:
                        continue
                    if e['supplier']:
                        cat = 'デッド' if e['dead'] else '期限切迫'
                        holders.append('%s（%s・数量%s・期限%s）'
                                       % (e['store'], cat, fmt_qty(e['over_qty']), fmt_ym(e['exp'])))
            shortage_rows.append({
                '店': s['name'], '薬品名': g(row, '薬品名'), '在庫数': round(zaiko, 2),
                '安全在庫数': round(safe, 2), '不足数': round(safe - zaiko, 2),
                '医薬品CD': key or '',
                'デッド/期限切迫で持つ他店': ' / '.join(holders) if holders else '（なし）'})
    shortage_rows.sort(key=lambda r: (r['店'], -r['不足数']))

    # --- 品目×店舗マトリクス ---
    #   行に出す品目＝出し手（デッド／期限切迫）がある品目。セルは：
    #     デX＝デッドで数量X ／ 限X＝期限切迫で数量X ／ 不＝安全在庫割れ ／ 使＝直近6ヶ月出庫あり
    over_keys = {}
    for (store, key), e in store_key_info.items():
        if e['supplier']:
            over_keys.setdefault(key, e['name'])
    matrix_keys = sorted(over_keys.keys(), key=lambda k: over_keys[k])
    store_names = [s['name'] for s in stores]
    matrix_rows = []
    for key in matrix_keys:
        cells = [over_keys[key], key]
        for sn in store_names:
            e = store_key_info.get((sn, key))
            if e is None:
                cells.append('')
            elif e['supplier']:
                # デッド優先（デッドかつ期限切迫なら「デ」）。純粋な期限切迫だけ「限」。
                mark = 'デ' if e['dead'] else '限'
                cells.append('%s%s' % (mark, fmt_qty(e['over_qty'])))
            elif e['shortage']:
                cells.append('不')
            elif e['usage_cnt6'] > 0:
                cells.append('使')
            else:
                cells.append('')
        matrix_rows.append(cells)

    # --- 店舗別サマリ ---
    #   出し手（デッド／期限切迫）を、排他・デッド優先で数える（proposal と同じ数え方）。
    #   金額は「在庫金額（在庫数×薬価）」を使う。麻薬・覚醒剤（法規制除外）は数えない。
    summary_rows = []
    for s in stores:
        dead_cnt = 0
        dead_amt = 0.0
        exp_cnt = 0
        exp_amt = 0.0
        short_cnt = 0
        for row in s['rows']:
            if (is_supplier(row, s['base_date']) and not is_legal_excluded(row)
                    and not is_excluded(s['name'], row)
                    and not is_below_min_amount(row)):
                amt = stock_amount(row)
                if supplier_category(row, s['base_date']) == 'デッド':
                    dead_cnt += 1
                    dead_amt += amt
                else:  # 期限切迫（デッドでない）
                    exp_cnt += 1
                    exp_amt += amt
            if is_shortage(row):
                short_cnt += 1
        summary_rows.append({
            '店': s['name'], 'デッド品目数': dead_cnt, 'デッド金額計': round(dead_amt, 2),
            '期限切迫品目数': exp_cnt, '期限切迫金額計': round(exp_amt, 2),
            '不足品目数': short_cnt})

    # --- 自己検算：出し手（デッド＋期限切迫・法規制除外後）の在庫金額合計 と 融通提案シートの合計 ---
    checkA = 0.0
    for s in stores:
        for row in s['rows']:
            if (is_supplier(row, s['base_date']) and not is_legal_excluded(row)
                    and not is_excluded(s['name'], row)
                    and not is_below_min_amount(row)):
                checkA += stock_amount(row)
    checkB = sum(r['_amt_raw'] for r in proposal_rows)
    diff = abs(checkA - checkB)
    check_ok = diff < 0.01

    return {
        'proposal_rows': proposal_rows,
        # ④受け手ビュー用：引取候補店を「1件1行」で持つ丸め無しの構造化データ（既存キーは一切変えず、増やすだけ）
        'candidate_rows': candidate_rows,
        # ④の別枠『いまは在庫があるが、先になら引き取れる薬』用（use_ref）。本来の④とは完全分離。
        #   ★Gシート／Excel には書き戻さない（結果4タブの payload には含めない）。
        'candidate_rows_ref': candidate_rows_ref,
        'shortage_rows': shortage_rows,
        'store_names': store_names,
        'matrix_keys': matrix_keys,
        'matrix_rows': matrix_rows,
        'summary_rows': summary_rows,
        'dist': dist,
        'nokey_rows': nokey_rows,
        'nokey_overstock': nokey_overstock,
        'legal_excluded_count': legal_excluded_count,
        'legal_excluded_amt': legal_excluded_amt,
        'user_excluded_count': user_excluded_count,
        'small_excluded_count': small_excluded_count,
        'small_excluded_amt': small_excluded_amt,
        'small_by_store': small_by_store,
        'min_supply_amount': CONFIG['min_supply_amount'],
        'checkA': checkA,
        'checkB': checkB,
        'diff': diff,
        'check_ok': check_ok,
    }


# ============================================================================
# 滞留（同じ品が何ヶ月つづけてリストに載っているか）
# ----------------------------------------------------------------------------
# 2026-07-30 追加。毎月アップロードしても動かない品を目立たせるための仕組み。
#
# 【何のため】
#   「引取候補店がいるのに、2ヶ月目・3ヶ月目と引き取られないまま残っている品」を
#   一目で分かるようにする（本間部長への指摘より）。
#
# 【色の意味】★画面（Streamlit）とExcelで必ず同じ色・同じ言葉を使うため、ここに一本化する。
#   ・行の色  … 期限のこと（既存。赤＝有効期限まで5ヶ月以内／黄＝期限1年以内）
#   ・滞留列の色 … 動いていない期間のこと（ここで定義）
#   この2つは別のことを表しているので、色の系統を分けている。
#
#   new     塗らない  今月から載った品（初めて出てきた）
#   m2      薄い黄    2ヶ月目。引取候補店はいるのに動いていない
#   m3      オレンジ  3ヶ月目。同上（引取候補店へ連絡を）
#   m4      濃い橙    4ヶ月目以上。同上（最優先で動かす）
#   nocand  グレー    2ヶ月目以上だが引取候補店がいない
#                     ＝引き取れる店がそもそも無い。店の努力では動かせないので、
#                       暖色で急かさずグレーにして区別する
#                       （社内で動かせない品なので、リバイバルドラッグなど社外の手段を検討する）。
#   booked  紫        先月ほかの店が予約したのに、今月もまだ残っている“のに、今月はもう有効な予約が無い”
#                     ＝約束したのに受け渡さず宙に浮いた品（取りこぼし）。最優先で確認する。
#   reserved 塗らない  いま有効な予約が生きている品（例：3ヶ月後に受け取る約束で待っている最中）。
#                     ★2026-08-04 追加。②の改修で予約が受取予定月まで持ち越されるようになり、
#                       約束どおり待っているだけの品まで booked（紫）で警告され続ける誤検知が出た。
#                       いま予約が生きている行は【塗らず・文字も出さない】。状態は出し手の
#                       『引取候補店』欄に「✅ ◯◯ が引取予定（受取：Nヶ月後）」と出ているので、
#                       滞留の欄で重ねて警告する必要はない（オレンジに落とすと催促になり不適切）。
#                     ★凡例には出さない（STAGNATION_LEGEND_ORDER に入れない）＝色も説明も無し。
#                       ただし _滞留月数 の数え上げは止めない（下の apply_stagnation / _滞留persist 参照）。
# ============================================================================

# 滞留区分 → (画面用の背景色, 画面用の文字色, Excelの塗り色, 凡例の説明)
#   ★文字色を黒で明示するのは、ダークテーマだと白文字×淡い背景で読めなくなるため
#     （既存の _style_expiry と同じ理由）。
STAGNATION_STYLES = {
    'new':      ('',        '',        None,     '今月から載った品'),
    'm2':       ('#FFF0B3', '#000000', 'FFF0B3', '2ヶ月目（引取候補店はいるのに動いていない）'),
    'm3':       ('#FFCB7A', '#000000', 'FFCB7A', '3ヶ月目（引取候補店へ連絡を）'),
    'm4':       ('#F2925E', '#000000', 'F2925E', '4ヶ月目以上（最優先で動かす）'),
    'nocand':   ('#E4E4E4', '#000000', 'E4E4E4',
                 '引き取れる店舗がありません（リバイバルドラッグなどを検討）'),
    'booked':   ('#D6B3E8', '#000000', 'D6B3E8', '先月予約されたのに残っている（受け渡し未完了）'),
    # 予約が生きている行＝塗らない・文字も出さない。凡例にも出さない（LEGEND_ORDER に入れない）。
    'reserved': ('',        '',        None,     '予約が入っており受け渡し待ち（塗らない）'),
}

# 凡例に出す順番（画面の「色の見方」で使う）。★'reserved' は入れない＝凡例に出さない。
STAGNATION_LEGEND_ORDER = ['m2', 'm3', 'm4', 'booked', 'nocand']


def stagnation_view(months, has_candidate, was_reserved, now_reserved=False):
    """
    滞留の状態から（区分, 画面に出す文字）を返す。

      months        … 何ヶ月つづけて載っているか（今月だけなら 1）
      has_candidate … 今月、引取候補店がいるか（True/False）
      was_reserved  … 先月、どこかの店が予約していたか（True/False）
      now_reserved  … 今月、いま有効な予約が生きているか（True/False）★2026-08-04 追加

    判定の優先順位（上が強い）：
      0. いま有効な予約が生きている      → reserved（塗らない・文字なし）★最優先
                                           約束どおり待っている最中なので警告しない。
      1. 先月予約されたのに残っている    → booked（＝先月予約あり かつ 今そのが予約は無い＝取りこぼし）
      2. 今月から載った品                → new（塗らない）
      3. 引取候補店がいない              → nocand（グレー。店のせいではない）
      4. 2/3/4ヶ月目以上                 → m2 / m3 / m4

    ★0 を 1 より前に置くのが今回の修正の肝。持ち越し予約は「先月も予約あり(was)」かつ
      「今も予約あり(now)」なので、now を先に見ないと booked（紫）に誤判定される。
      逆に「先月は予約あり・今は予約が消えている」＝ was だけ True のときが本物の取りこぼし＝booked。
    """
    try:
        months = int(months)
    except (TypeError, ValueError):
        months = 1
    if months < 1:
        months = 1

    # 0. いま有効な予約が生きている品は塗らない・文字も出さない（約束どおり進行中）
    if now_reserved:
        return 'reserved', ''

    # 1. 先月予約されたのに、今月はもう有効な予約が無い＝受け渡しの取りこぼし
    if was_reserved:
        return 'booked', '⚠ %dヶ月目・先月予約済' % months

    # 2. 今月から載った品は塗らない（毎月ほとんどの行が新規なので、塗ると意味が薄れる）
    if months <= 1:
        return 'new', ''

    # 3. 引き取れる店がそもそもいない品は、急かしても動かないのでグレーで区別する
    if not has_candidate:
        return 'nocand', '%dヶ月目（候補店なし）' % months

    # 4. 候補店がいるのに動いていない＝今回いちばん見せたいところ
    if months == 2:
        return 'm2', '2ヶ月目'
    if months == 3:
        return 'm3', '3ヶ月目'
    return 'm4', '%dヶ月目' % months


# ============================================================================
# Excel出力（コマンド版のExcelダウンロードと、アプリ版のExcelダウンロードで共用）
# ============================================================================
RED_FILL = PatternFill('solid', fgColor='FFC7CE')     # 期限切迫＝有効期限まで5ヶ月以内（赤系）
DEAD_FILL = PatternFill('solid', fgColor='FCE4D6')    # デッド（薄オレンジ系）※マトリクスの「デX」用
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')  # 期限1年以内（黄系）
HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')   # 見出し
NOTE_FILL = PatternFill('solid', fgColor='FFF2CC')     # 注記
REF_FILL = PatternFill('solid', fgColor='EDEDED')      # 参考(過剰だが使用中)＝本命と区別する薄グレー
REF_HEADER_FILL = PatternFill('solid', fgColor='D0CECE')  # 参考列の見出し（濃いめグレー）
HEADER_FONT = Font(bold=True)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _auto_width(ws, headers, max_width=48):
    """ 列幅をざっくり自動調整 """
    for ci, h in enumerate(headers, start=1):
        col = get_column_letter(ci)
        maxlen = len(str(h))
        for cell in ws[col]:
            if cell.value is not None:
                # 全角をやや広めに数える
                s = str(cell.value)
                w = sum(2 if ord(ch) > 0x2E7F else 1 for ch in s)
                maxlen = max(maxlen, w)
        ws.column_dimensions[col].width = min(maxlen + 2, max_width)


def write_excel(path, base_ym_disp, csv_base_disp,
                proposal_rows, shortage_rows,
                matrix_store_names, matrix_rows,
                summary_rows):
    wb = Workbook()

    # ---------- シート1：融通提案 ----------
    ws = wb.active
    ws.title = '融通提案'
    note = ('基準年月：%s ／ CSV基準日：%s ／ '
            '※これは月初スナップショットであり、現在庫と異なる場合があります。'
            % (base_ym_disp, csv_base_disp))
    ws.cell(row=1, column=1, value=note).fill = NOTE_FILL
    ws.cell(row=1, column=1).font = Font(bold=True, color='7F6000')

    # ※末尾の『予約』は2026-07-28に追加した列（受け手の店が押さえた印）。
    #   ※末尾の『滞留』は2026-07-30に追加した列（何ヶ月つづけて載っているか）。
    #   既存の列は順番も中身も変えていないので、旧版との照合は追加2列を除いて行える。
    #   ★新しい列を末尾に足すのは、途中に差し込むと過去のExcelとの列位置が合わなくなるため。
    headers = ['出し手店', '種別', '薬品名', '単位', 'メーカ名', '在庫数', '在庫金額',
               '過剰在庫区分', '不動区分', '期限切迫区分', '有効期限', 'ロットNO',
               '最終出庫日', '区分', '引取候補店', '参考:過剰だが使用中の店',
               '6ヶ月出庫回数', '医薬品CD', '予約', '滞留']
    # 「参考」列は本命（引取候補店）とはっきり区別するため、見出し・セルをグレーで塗る
    ref_col_idx = headers.index('参考:過剰だが使用中の店') + 1  # 1始まり
    # 「滞留」列は行の色（＝期限のこと）とは別に、滞留区分ごとの色で塗る
    stag_col_idx = headers.index('滞留') + 1  # 1始まり
    hr = 2
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.fill = REF_HEADER_FILL if ci == ref_col_idx else HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
    r = hr + 1
    for pr in proposal_rows:
        vals = [pr['出し手店'], pr['種別'], pr['薬品名'], pr['単位'], pr['メーカ名'], pr['在庫数'],
                pr['在庫金額'], pr['過剰在庫区分'], pr['不動区分'], pr['期限切迫区分'],
                pr['有効期限'], pr['ロットNO'], pr['最終出庫日'], pr['区分'],
                pr['引取候補店'], pr['参考:過剰だが使用中の店'],
                pr['6ヶ月出庫回数'], pr['医薬品CD'], pr.get('予約', ''),
                pr.get('滞留', '')]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = BORDER
            # 在庫数（6列目）・在庫金額（7列目）は小数第2位まで表示する
            if ci in (6, 7):
                cell.number_format = '#,##0.00'
        # 塗り分け：有効期限まで5ヶ月以内は赤、そうでなく1年以内は黄（ただし「参考」列は対象外）
        fill = None
        if pr['_expiry_flag']:
            fill = RED_FILL
        elif pr['_remain'] is not None and pr['_remain'] <= CONFIG['expiry_yellow_months']:
            fill = YELLOW_FILL
        if fill:
            for ci in range(1, len(headers) + 1):
                if ci in (ref_col_idx, stag_col_idx):
                    continue
                ws.cell(row=r, column=ci).fill = fill
        # 「参考」列は行の色に関係なく常にグレー＝本命候補と一目で区別できるようにする
        ws.cell(row=r, column=ref_col_idx).fill = REF_FILL
        # 「滞留」列は行の色（期限のこと）とは別系統の色で塗る。
        #   行の赤／黄に上書きされないよう、上の塗りつぶしから除外している。
        stag_color = STAGNATION_STYLES.get(pr.get('_滞留区分', 'new'), (None, None, None, ''))[2]
        if stag_color:
            ws.cell(row=r, column=stag_col_idx).fill = PatternFill('solid', fgColor=stag_color)
        r += 1
    ws.auto_filter.ref = '%s%d:%s%d' % ('A', hr, get_column_letter(len(headers)), max(hr, r - 1))
    ws.freeze_panes = 'A%d' % (hr + 1)
    _auto_width(ws, headers)

    # ---------- シート2：不足品目一覧 ----------
    ws2 = wb.create_sheet('不足品目一覧')
    headers2 = ['店', '薬品名', '在庫数', '安全在庫数', '不足数', '医薬品CD', 'デッド/期限切迫で持つ他店']
    for ci, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
    r = 2
    for sr in shortage_rows:
        vals = [sr['店'], sr['薬品名'], sr['在庫数'], sr['安全在庫数'], sr['不足数'],
                sr['医薬品CD'], sr['デッド/期限切迫で持つ他店']]
        for ci, v in enumerate(vals, start=1):
            ws2.cell(row=r, column=ci, value=v).border = BORDER
        r += 1
    ws2.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(headers2)), max(1, r - 1))
    ws2.freeze_panes = 'A2'
    _auto_width(ws2, headers2)

    # ---------- シート3：品目×店舗マトリクス ----------
    ws3 = wb.create_sheet('品目×店舗マトリクス')
    headers3 = ['薬品名', '医薬品CD'] + matrix_store_names
    for ci, h in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
    r = 2
    for mrow in matrix_rows:
        for ci, v in enumerate(mrow, start=1):
            cell = ws3.cell(row=r, column=ci, value=v)
            cell.border = BORDER
            if ci >= 3 and isinstance(v, str):
                if v.startswith('限'):
                    cell.fill = RED_FILL       # 期限切迫（赤）
                elif v.startswith('デ'):
                    cell.fill = DEAD_FILL      # デッド（薄オレンジ）
                elif v == '不':
                    cell.fill = YELLOW_FILL
        r += 1
    # 凡例
    ws3.cell(row=r + 1, column=1,
             value='凡例：デX=デッド（6ヶ月以上出庫なし・数量X） ／ 限X=期限切迫（有効期限まで5ヶ月以内・数量X） ／ '
                   '不=安全在庫割れ ／ 使=直近6ヶ月に出庫実績あり ／ 空欄=在庫なし')
    ws3.freeze_panes = 'C2'
    _auto_width(ws3, headers3, max_width=14)
    ws3.column_dimensions['A'].width = 32

    # ---------- シート4：店舗別サマリ ----------
    ws4 = wb.create_sheet('店舗別サマリ')
    headers4 = ['店', 'デッド品目数', 'デッド金額計', '期限切迫品目数', '期限切迫金額計', '不足品目数']
    for ci, h in enumerate(headers4, start=1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
    r = 2
    for su in summary_rows:
        vals = [su['店'], su['デッド品目数'], su['デッド金額計'], su['期限切迫品目数'],
                su['期限切迫金額計'], su['不足品目数']]
        for ci, v in enumerate(vals, start=1):
            cell = ws4.cell(row=r, column=ci, value=v)
            cell.border = BORDER
            # デッド金額計（3列目）・期限切迫金額計（5列目）は小数第2位まで
            if ci in (3, 5):
                cell.number_format = '#,##0.00'
        r += 1
    ws4.freeze_panes = 'A2'
    _auto_width(ws4, headers4)

    wb.save(path)


# ============================================================================
# 引取依頼書（帳票）Excel ＝ 受け手（引き取る側）が予約した品を、出し手店ごとの
#   シートにまとめた「もらいに行くための紙」。デスクネッツ貼付・FAX・電話しながらの参照用。
# ----------------------------------------------------------------------------
# 2026-08-01 追加。★write_excel（全店4シート・画面/分析用）とは別物です。
#   ・write_excel は絶対に無編集（品質管理部が毎回「全セル差分ゼロ」で回帰を取っている）。
#   ・こちらは新設。印刷設定（A4横・横1ページ）を入れるのはこちらのシートだけ。
#     write_excel に印刷設定を足すと回帰の前提が変わるため、あちらには入れない。
#
# ★FAX（白黒）運用の判断（本間部長 追加指示 2026-08-01）
#   FAXは白黒なので、期限切迫の赤（RED_FILL）や1年以内の黄（YELLOW_FILL）で行を塗ると、
#   送った先ではどちらも同じ灰色の塊になり、かえって字が読みにくくなります。
#   そこで【この帳票では期限を色で塗らず】、期限が近い品は『有効期限』欄の文字を太字にして
#   伝えます（塗りに意味を持たせない）。見出し行だけは薄い色（HEADER_FILL）で、
#   白黒でも「ここが列名」と分かる程度に留めます。
#   （既存の小道具 HEADER_FILL / BORDER / HEADER_FONT / _auto_width は呼ぶだけで使い回し、
#     RED_FILL / YELLOW_FILL は意味づけには使いません。）
# ============================================================================

# 引取依頼書の明細列（この順・本間部長確定）。在庫金額は載せない（発注用の紙なので）。
PICKUP_REQUEST_COLS = ['薬品名', '単位', '数量', '有効期限', 'ロットNO', '医薬品CD',
                       '受取予定月', '区分', '状態']

# 各列の幅（文字数）。★_auto_width の結果をそのまま使わず、A4横に収まる上限を決め打ちする。
#   合計が A4横の印刷可能幅（余白を引いた実寸）に収まることを、生成後に数値で検算する
#   （過去に図の表で右端が枠外に切れた事故があり、列幅合計を枠内に収める検算で解決した）。
PICKUP_REQUEST_WIDTHS = {
    '薬品名': 28, '単位': 6, '数量': 8, '有効期限': 12, 'ロットNO': 14,
    '医薬品CD': 14, '受取予定月': 16, '区分': 10, '状態': 22,
}

# Excelのシート名に使えない文字。将来の飛鳥22店追加に備えた保険（現行14店では発動しない）。
_SHEET_NAME_BAD_CHARS = set('[]:*?/\\')


def _safe_sheet_name(name):
    """ Excelのシート名として安全な名前にする：禁止文字 []:*?/\\ を除き、31文字以内に切る。
        ★現行14店（東大泉／海浜幕張／…／下落合）はすべて禁止文字なし・31文字以内で、この関数は
          実質何もしない。将来 飛鳥22店などを足したときに黙って壊れないための保険。 """
    s = ''.join(ch for ch in str(name) if ch not in _SHEET_NAME_BAD_CHARS).strip()
    if not s:
        s = 'シート'
    return s[:31]


def write_pickup_request_excel(path_or_buf, data):
    """ 引取依頼書Excelを書き出す（FAX・デスクネッツ用）。
        data … app_logic.build_pickup_request の戻り：
          {'my_store': 自店名, 'ym': 'YYYYMM',
           'sheets': [{'出し手店': 店名,
                       'rows': [{'薬品名','単位','数量','有効期限','ロットNO','医薬品CD',
                                 '受取予定月','区分','状態','_期限強調'(bool)}, ...]}, ...]}
        ・シート＝出し手店ごと（さと和光シート／東立石シート…）。シート名は出し手店名そのまま。
        ・PDFは作らない（Excel 1ファイル）。ファイル名は呼び出し側で付ける。 """
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = Workbook()
    wb.remove(wb.active)   # 既定の空シートは消し、出し手店ごとに作り直す

    my_store = data.get('my_store', '')
    ym = data.get('ym', '')
    ym_disp = ('%s年%s月' % (ym[:4], ym[4:6])) if (ym and len(str(ym)) >= 6) else '不明'
    today = datetime.date.today().strftime('%Y/%m/%d')

    ncol = len(PICKUP_REQUEST_COLS)
    last_col = get_column_letter(ncol)
    # FAXで潰れないよう本文は10pt（9pt以上の目安を満たす）。太字は期限強調用。
    body_font = Font(size=10)
    body_bold = Font(size=10, bold=True)

    used_names = set()
    for sheet in data.get('sheets', []):
        # --- シート名（禁止文字除去・31文字・重複回避）---
        title = _safe_sheet_name(sheet.get('出し手店', ''))
        base = title
        n = 2
        while title in used_names:                 # 万一同名になったら連番（現行14店では起きない）
            title = '%s%d' % (base[:28], n)
            n += 1
        used_names.add(title)
        ws = wb.create_sheet(title=title)

        # --- 上部ヘッダー欄（罫線付き・横幅いっぱいに結合）---
        r = 1
        tcell = ws.cell(row=r, column=1, value='引取依頼書')
        tcell.font = Font(size=14, bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        r += 1
        for text in ['依頼元（引き取る店）：%s' % my_store,
                     '出し手店（もらう先）：%s' % sheet.get('出し手店', ''),
                     '作成日：%s' % today,
                     '対象年月：%s' % ym_disp]:
            cell = ws.cell(row=r, column=1, value=text)
            cell.font = body_bold
            cell.border = BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            r += 1
        for text in ['数量は在庫まるごとを初期値にしています。減らす場合はこの欄を手で書き換えてください。',
                     '『区分』に表示のある薬（向精神薬・毒薬・劇薬）は、受け取る側でも譲受の記録が必要です。']:
            cell = ws.cell(row=r, column=1, value=text)
            cell.font = body_font
            cell.fill = NOTE_FILL   # 薄い注記色（白黒FAXでもうっすら残る程度で、字は読める）
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            r += 1
        r += 1   # ヘッダーと明細のあいだを1行あける

        # --- 明細ヘッダー（この行を全ページで繰り返す）---
        header_row = r
        for ci, h in enumerate(PICKUP_REQUEST_COLS, start=1):
            hc = ws.cell(row=r, column=ci, value=h)
            hc.fill = HEADER_FILL          # 見出しだけは薄色（白黒でも列名と分かる程度）
            hc.font = HEADER_FONT
            hc.border = BORDER
            hc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        r += 1
        first_detail = r

        # --- 明細 ---
        for d in sheet.get('rows', []):
            for ci, colname in enumerate(PICKUP_REQUEST_COLS, start=1):
                cell = ws.cell(row=r, column=ci, value=d.get(colname, ''))
                cell.border = BORDER
                # 期限が近い品は【塗らずに】有効期限の文字だけ太字（白黒FAX対策）
                if colname == '有効期限' and d.get('_期限強調'):
                    cell.font = body_bold
                else:
                    cell.font = body_font
                if colname in ('薬品名', '状態'):
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='top')
            r += 1
        last_detail = max(r - 1, header_row)

        # --- 列幅（A4横に収まる上限つき固定。_auto_width は使わない）---
        for ci, h in enumerate(PICKUP_REQUEST_COLS, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = PICKUP_REQUEST_WIDTHS[h]

        # --- 印刷設定（A4横・横は必ず1ページ・縦は品数に応じて何ページでも）---
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        # ★fitToPage=True を立てないと fitToWidth は効かない
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                      header=0.2, footer=0.2)
        # 明細の列名行を全ページで繰り返す（2ページ目以降で列名の無い紙が出ないように）
        ws.print_title_rows = '%d:%d' % (header_row, header_row)
        ws.print_area = 'A1:%s%d' % (last_col, last_detail)
        # フッター右にページ番号（FAXで枚数の取り違えを防ぐ）
        ws.oddFooter.right.text = '&P / &N'
        ws.freeze_panes = 'A%d' % first_detail

    # 予約が1件も無い（シート0枚）の保険。呼び出し側でボタンを出さない想定だが、空ブック回避。
    if not wb.sheetnames:
        ws = wb.create_sheet('引取依頼書')
        ws.cell(row=1, column=1, value='予約されている品がありません。')

    wb.save(path_or_buf)
